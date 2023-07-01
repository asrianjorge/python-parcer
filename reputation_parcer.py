import datetime
import json
import random
import time
from copy import copy

from bs4 import BeautifulSoup
import requests
import html5lib
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Font
import lxml

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36'
}

border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))


def yandex(URL):
    if URL == "0":
        return "0"
    else:
        response = requests.get(URL, headers=HEADERS)
        soup = BeautifulSoup(response.text, 'lxml')

        try:
            range = soup.find('span', class_='business-rating-badge-view__rating-text _size_m').text.strip()
        except AttributeError:
            range = 'error'

        try:
            reviews = soup.find('span', class_='business-header-rating-view__text _clickable').text.strip().split(' ')[
                0]
        except AttributeError:
            reviews = 'error'

        try:
            ratings = soup.find('span', class_='business-rating-amount-view _summary').text.strip().split(' ')[0]
        except AttributeError:
            ratings = 'error'

        return f"{range} ({reviews}) ({ratings})"


def google(URL):
    if URL == "0":
        return "0"
    else:
        response = requests.get(URL, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'lxml')

        # try:
        #     range = soup.find('div', class_='gm2-display-2').text.strip()
        # except AttributeError:
        #     range = 'error'
        #
        # try:
        #     reviews = soup.find('div', class_='gm2-caption').text.strip().split(' ')[1]
        # except AttributeError:
        #     reviews = 'error'

        try:
            range = soup.find('span', class_='aMPvhf-fI6EEc-KVuj8d').text.strip()
        except AttributeError:
            range = 'error'

        # try:
        #     reviews = soup.find('button', class_='widget-pane-link').text.strip().split(' ')[0]
        # except AttributeError:
        #     reviews = 'error'

        try:
            reviews = soup.find('span', class_='OAO0-ZEhYpd-vJ7A6b OAO0-ZEhYpd-vJ7A6b-qnnXGd').text.strip().split(' ')[0]
        except AttributeError:
            reviews = 'error'

        return f"{range} ({reviews})"


def zoon(URL):
    if URL == "0":
        return "0"
    else:
        response = requests.get(URL, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'lxml')
        try:
            range = soup.find('span', class_='rating-value').get_text(strip=True)
        except AttributeError:
            range = 'error'
        try:
            reviews = soup.find('a', class_='fs-large gray js-toggle-content').get_text(strip=True).split(' ')[
                0].strip()
        except AttributeError:
            reviews = 'error'
        try:
            ratings = soup.find('span', class_='js-total-count').get_text(strip=True).strip()
        except AttributeError:
            ratings = 'error'

        return f"{range} ({reviews}) ({ratings})"


def gis(URL):
    if URL == "0":
        return "0"
    else:
        response = requests.get(URL, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'lxml')

        try:
            range = soup.find('span', class_='_1n8h0vx').get_text(strip=True)
        except AttributeError:
            range = 'error'

        try:
            reviews = soup.find('span', class_='_gg5kmr').get_text(strip=True).split(' ')[0].strip()
        except AttributeError:
            reviews = 'error'

        # try:
        #     range = soup.find('span', class_='_1n8h0vx').get_text(strip=True)
        # except AttributeError:
        #     range = 'error'
        #
        # try:
        #     reviews = soup.find('a', class_='_gg5kmr').get_text(strip=True).split(' ')[0].strip()
        # except AttributeError:
        #     reviews = 'error'

        return f"{range} ({reviews})"


def trip(URL):
    if URL == "0":
        return "0"
    else:
        response = requests.get(URL, headers=HEADERS)
        soup = BeautifulSoup(response.content, 'lxml')

        try:
            range = soup.find('div', class_='WlYyy cPsXC fksET cMKSg').get_text(strip=True)
        except AttributeError:
            range = 'error'

        try:
            reviews = soup.find('span', class_='WlYyy diXIH bGusc dDKKM').get_text(strip=True).split(' ')[0].strip()
        except AttributeError:
            reviews = 'error'

        # try:
        #     range = soup.find('div', class_='DrjyGw-P _1SRa-qNz _3t0zrF_f _1QGef_ZJ').get_text(strip=True)
        # except AttributeError:
        #     range = 'error'
        #
        # try:
        #     reviews = soup.find('span', class_='DrjyGw-P _26S7gyB4 _14_buatE _2nPM5Opx').get_text(strip=True).split(' ')[0].strip()
        # except AttributeError:
        #     reviews = 'error'

        return f"{range} ({reviews})"


def borderize(sheet, write, col):
    sheet[write][col].border = Border(left=Side(border_style='thin', color='FF000000'),
                                      right=Side(border_style='thin', color='FF000000'),
                                      top=Side(border_style='thin', color='FF000000'),
                                      bottom=Side(border_style='thin', color='FF000000'))


def colorize(sheet, watch, write, col):
    try:
        a = sheet[watch][col].value.split(' ')[0]
    except AttributeError:
        a = sheet[watch][col].value

    try:
        b = sheet[write][col].value.split(' ')[0]
    except AttributeError:
        b = sheet[write][col].value

    # print(a, b)

    try:
        a = '.'.join(a.split(','))
    except AttributeError:
        a = a
    except ValueError:
        a = a

    try:
        b = '.'.join(b.split(','))
    except AttributeError:
        b = b
    except ValueError:
        b = b

    try:
        if b == 0 or float(b) == 0.0:
            sheet[write][col].fill = PatternFill(fill_type='solid', start_color='ffffff')
        elif a is not None:
            if float(a) < float(b):
                sheet[write][col].fill = PatternFill(fill_type='solid', start_color='00ff00')
            elif float(a) == float(b) and float(b) == float('5.0'):
                sheet[write][col].fill = PatternFill(fill_type='solid', start_color='b6d7a8')
            elif float(a) > float(b):
                sheet[write][col].fill = PatternFill(fill_type='solid', start_color='ff0000')
            elif float(a) == float(b):
                sheet[write][col].fill = PatternFill(fill_type='solid', start_color='ffff00')
    except ValueError:
        # sheet[write][col].fill = PatternFill(fill_type='solid', start_color='9C00FF')
        sheet[write][col].fill = PatternFill(fill_type='solid', start_color='ffffff')

    # making bolds
    try:
        if b != 0 or float(b) != 0.0:
            sheet[write][col].font = Font(bold=False)
        elif float(b) < float(4.3):
            sheet[write][col].font = Font(bold=True)
        else:
            sheet[write][col].font = Font(bold=False)
    except ValueError:
        sheet[write][col].font = Font(bold=False)

    borderize(sheet, write, col)
    print(b)


def add_to_table(mass):
    data = mass

    book = openpyxl.open('MissionReputation.xlsx', read_only=False)
    sheet = book['Лист1']

    # Обозначим все необходимые числа
    count = sheet['K2'].value  # кол-во выполненных миссий
    k = 30  # коэффициент
    mini_k = 7
    arg = 0
    watching_row = (count - 1) * k + 1
    writing_row = count * k + 1  # оспоставляем данные с watching и записываем в writing

    sheet[writing_row][0].value = datetime.date.today()

    # sheet['A1'].value = datetime.date.today()
    # sheet['A2'].value = 'Яндекс'
    # sheet['A2'].value = 'Гугл'
    # sheet['A2'].value = 'Zoon'
    # sheet['A2'].value = '2gis'
    # sheet['A2'].value = 'Tripadvisor'

    for col in range(0, len(data)):
        if ("next" in data[col][0]) or (col - arg == 0):
            if col != 0:
                watching_row += mini_k
                writing_row += mini_k
            arg = col

            sheet[writing_row + 1][col - arg].value = 'Яндекс'
            sheet[writing_row + 2][col - arg].value = 'Гугл'
            sheet[writing_row + 3][col - arg].value = 'Zoon'
            sheet[writing_row + 4][col - arg].value = '2gis'
            sheet[writing_row + 5][col - arg].value = 'Tripadvisor'

            borderize(sheet, writing_row + 1, col - arg)
            borderize(sheet, writing_row + 2, col - arg)
            borderize(sheet, writing_row + 3, col - arg)
            borderize(sheet, writing_row + 4, col - arg)
            borderize(sheet, writing_row + 5, col - arg)

            continue
        elif "next" not in data[col][0]:
            sheet[writing_row][col - arg].value = data[col][0]
            sheet[writing_row + 1][col - arg].value = data[col][1]
            sheet[writing_row + 2][col - arg].value = data[col][2]
            sheet[writing_row + 3][col - arg].value = data[col][3]
            sheet[writing_row + 4][col - arg].value = data[col][4]
            sheet[writing_row + 5][col - arg].value = data[col][5]
            # color it!
            sheet[writing_row][col - arg].fill = copy(sheet[watching_row][col - arg].fill)
            # sheet[writing_row][col - arg].fill = sheet[watching_row][col - arg].fill
            borderize(sheet, writing_row, col - arg)
            print(data[col][0])
            colorize(sheet, watching_row + 1, writing_row + 1, col - arg)
            colorize(sheet, watching_row + 2, writing_row + 2, col - arg)
            colorize(sheet, watching_row + 3, writing_row + 3, col - arg)
            colorize(sheet, watching_row + 4, writing_row + 4, col - arg)
            colorize(sheet, watching_row + 5, writing_row + 5, col - arg)
        print(col, arg, col - arg)

    sheet['K2'].value += 1
    book.save(filename='MissionReputation.xlsx')


def parcer(filename):
    # Создаем словарь с данными из json файла
    with open(filename, 'r', encoding='utf-8') as file:
        businesses = json.load(file)

    # Создаем массив с данными о всех предприятиях, и добавляем эти данные
    all = []
    for i in businesses:
        all.append((i,
                    yandex(businesses[i][0]),
                    google(businesses[i][1]),
                    zoon(businesses[i][2]),
                    gis(businesses[i][3]),
                    trip(businesses[i][4])
                    ))
        print(i + ' done')

    # Добавляем данные в таблицу
    add_to_table(all)

    # print(*all)


parcer('for_parcing.json')
# print(gis("https://2gis.ru/spb/firm/70000001036249801"))
# print(trip("https://www.tripadvisor.ru/Attraction_Review-g298484-d17380617-Reviews-Culinary_Studio_Igra_Stolov-Moscow_Central_Russia.html"))
# print(google("https://www.google.com/maps/place/Арт+Школа+№1/@59.9206361,30.3543966,17z/data=!4m5!3m4!1s0x0:0x490fce03cbbdfa6a!8m2!3d59.9206361!4d30.3565853"))
# print(google("https://maps.google.com/maps?cid=13527272279089820521&_ga=2.32885054.1028217334.1631530778-787220362.1625669438"))